import h5py
import numpy as np
import openpyxl

from sklearn.preprocessing import OneHotEncoder

def get_Temperature(filename):
    wb = openpyxl.load_workbook(filename)
    sheet1 = wb['sheet1']
    Temperature=[]
    start=24.6
    for i in range(3420,2128-1,-1):
        value=float(sheet1.cell(i, 2).value)
        mid=(start+value)/2.0
        mid=format(mid,".1f")
        for j in range(3):
            Temperature.append(mid)
        for j in range(3):
            Temperature.append(value)
        start=value
        if i==9266:
            Temperature.append(value)

    return Temperature[:7220]

def get_Weather(filename):
    wb = openpyxl.load_workbook(filename)
    sheet1 = wb['2015']
    Weather=[]
    for i in range(306,367):
        str = sheet1.cell(i, 2).value
        str=str.split("/")
        for j in range(24):
            Weather.append(str[0])
        for j in range(24):
            Weather.append(str[1])

    sheet1 = wb['2016']
    for i in range(2,103):
        str = sheet1.cell(i, 2).value
        str=str.split("/")
        for j in range(24):
            Weather.append(str[0])
        for j in range(24):
            Weather.append(str[1])

    return Weather[:7220]

def get_holiday(filename):
    import datetime
    d1=datetime.datetime(2015,11,1)
    Holiday=[0 for i in range(7220)]
    data=[]

    for line in open(filename,"r"):
        if "\n" in line:
            line=int(line[:-1])
        else :
            line=int(line)
        year=int(format(line/10000,'.0f'))
        month=int(format((line%10000)/100,'2.0f'))
        day=int(format((line%100),'.0f'))

        if year <=2015 and month<=11:
            continue
        d2=datetime.datetime(year,month,day)
        interval = d2 - d1
        k=interval.days
        if((k+1)*48)>= len(Holiday):
            break
        for j in range(48):
            Holiday[k*48+j]=1

    return Holiday

def get_weekend():
    weekend = [0 for i in range(7220)]
    weekend[0]=1
    for i in range(1,7220):
        if i%7==6 or i%7==0:
            weekend[i]=1

    return weekend


def process_IF(dir):

    temperature=get_Temperature(dir+"北京海淀区气象站时间间隔3小时.xlsx")
    weather=get_Weather(dir+"北京天气爬虫.xlsx")
    holiday=get_holiday(dir+"BJ_Holiday.txt")
    weekend=get_weekend()
    #weather=set(weather)
    temperature = np.array(temperature).reshape(-1,1).astype( np.float32)
    holiday = np.array(holiday).reshape(-1, 1).astype( np.float32)
    weekend = np.array(weekend).reshape(-1, 1).astype( np.float32)
    weather = np.array(weather).reshape(-1,1)

    one_hot=OneHotEncoder()    
    weather=one_hot.fit_transform(weather).toarray().astype( np.float32)

    IF=np.concatenate((weather, holiday, weekend,temperature), axis=1)

    return temperature

def process(filename,s,m,l):
    IF = process_IF()

    f = h5py.File(filename)
    for ke in f.keys():
        print(ke, f[ke].shape)
    short=[]
    medium=[]
    long=[]
    object=[]
    data=f["data"][:,1,:,:]

    train_data=data[:-336,:,:]
    test_data=data[-336:,:,:]
    train_IF=IF[:-336,:]

    IF0=[]

    len=train_data.shape[0]
    #data processing
    for i in range(len):
        short_ele=[]
        medium_ele=[]
        long_ele=[]
        if (i + l * 48 * 7+1 > len):
            break
        for j in range(s):
            long_ele.append(train_data[i+j, :, :])
        for j in range(m):
            medium_ele.append(train_data[i+j*48, :, :])
        for j in range(l):
            short_ele.append(train_data[i+j*48*7,:,:])

        short.append(short_ele)
        medium.append(medium_ele)
        long.append(long_ele)
        object.append(train_data[i + j * 48 * 7 + 1, :, :])
        IF0.append(train_IF[i + j * 48 * 7 + 1, :])
    data = {}

    short=np.array(short)
    medium=np.array(medium)
    long =np.array(long)
    object = np.array(object)
    IF0=np.array(IF0)

    data["IF"]=IF0



    return  data

# data=process("TaxiBJ/BJ16_M32x32_T30_InOut.h5",s=2,m=2,l=2)
#
# np.save('TaxiBJ16/data_dict 二通道.npy', data) 













